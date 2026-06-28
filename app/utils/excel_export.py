"""
Excel export utility — builds formatted .xlsx workbooks from tabular data.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=10.5)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1B2A4A")


def build_excel_report(title, headers, rows, sheet_name="Report"):
    """
    headers: list of column header strings
    rows: list of lists/tuples matching header order
    Returns BytesIO buffer.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT

    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows]) if rows else len(str(header))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_csv_report(headers, rows):
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    byte_buf = io.BytesIO(buf.getvalue().encode("utf-8"))
    return byte_buf
